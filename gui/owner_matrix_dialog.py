from PySide6.QtWidgets import (
    QDialog,
    QVBoxLayout,
    QHBoxLayout,
    QTableWidget,
    QTableWidgetItem,
    QCheckBox,
    QPushButton,
    QMessageBox,
    QFileDialog,
    QTextEdit,
    QDialog,
    QAbstractItemView
)

from openpyxl import Workbook


class OwnerMatrixDialog(QDialog):

    def __init__(
        self,
        matrix,
        parent=None
    ):
        super().__init__(parent)

        self.matrix = matrix

        

        print(
            "MATRIX ROWS:",
            len(self.matrix)
        )

        self.setWindowTitle(
            "Owner Matrix"
        )

        self.resize(
            1000,
            600
        )

        layout = QVBoxLayout()

        filter_bar = QHBoxLayout()

        self.chk_hide_rows = QCheckBox(
            "Hide Empty Rows"
        )

        self.chk_hide_cols = QCheckBox(
            "Hide Empty Columns"
        )

        self.btn_export = QPushButton(
            "Export Excel"
        )

        filter_bar.addWidget(
            self.chk_hide_rows
        )

        filter_bar.addWidget(
            self.chk_hide_cols
        )

        filter_bar.addWidget(
            self.btn_export
        )

        layout.addLayout(
            filter_bar
        )

        self.chk_hide_rows.stateChanged.connect(
            self.refresh_matrix
        )

        self.chk_hide_cols.stateChanged.connect(
            self.refresh_matrix
        )

        self.btn_export.clicked.connect(
            self.export_excel
        )

        self.table = QTableWidget()
        
        self.table.setEditTriggers(
            QAbstractItemView.NoEditTriggers
        )

        self.table.setSelectionBehavior(
            QAbstractItemView.SelectItems
        )
       
        self.table.cellDoubleClicked.connect(
            self.show_owner_summary
        )
        
        self.owners = sorted(
            self.matrix.keys()
        )

        khewats = set()

        for data in self.matrix.values():

            for khewat in data.keys():

                khewats.add(
                    str(khewat)
                )

        self.khewats = sorted(
            list(khewats)
        )

        self.table.setRowCount(
            len(self.owners)
        )

        self.table.setColumnCount(
            len(self.khewats)
        )

        self.table.setVerticalHeaderLabels(
            self.owners
        )

        self.table.setHorizontalHeaderLabels(
            self.khewats
        )

        print(
            "ROWS:",
            len(self.owners)
        )

        print(
            "COLS:",
            len(self.khewats)
        )

        for row, owner in enumerate(
            self.owners
        ):

            owner_data = self.matrix.get(
                owner,
                {}
            )

            for col, khewat in enumerate(
                self.khewats
            ):

                value = owner_data.get(
                    khewat,
                    ""
                )

                self.table.setItem(
                    row,
                    col,
                    QTableWidgetItem(
                        str(value)
                    )
                )

        layout.addWidget(
            self.table
        )
        self.table.resizeColumnsToContents()
        
        self.setLayout(
            layout
        )

    def refresh_matrix(self):
        """
        Placeholder.
        Actual hide-row/hide-column logic
        can be implemented later.
        """
        pass

    def export_excel(self):

        file_name, _ = QFileDialog.getSaveFileName(
            self,
            "Export Owner Matrix",
            "owner_matrix.xlsx",
            "Excel Files (*.xlsx)"
        )

        if not file_name:
            return

        wb = Workbook()

        ws = wb.active

        ws.title = "Owner Matrix"

        # Header row

        ws.cell(
            row=1,
            column=1,
            value="Owner"
        )

        for col, khewat in enumerate(
            self.khewats,
            start=2
        ):

            ws.cell(
                row=1,
                column=col,
                value=str(khewat)
            )

        # Matrix data

        for row, owner in enumerate(
            self.owners,
            start=2
        ):

            ws.cell(
                row=row,
                column=1,
                value=owner
            )

            for col, khewat in enumerate(
                self.khewats,
                start=2
            ):

                item = self.table.item(
                    row - 2,
                    col - 2
                )

                value = ""

                if item:
                    value = item.text()

                ws.cell(
                    row=row,
                    column=col,
                    value=value
                )

        wb.save(
            file_name
        )

        QMessageBox.information(
            self,
            "Export Complete",
            f"Excel file saved:\n{file_name}"
        )

    def show_owner_summary(
        self,
        row,
        col
    ):

        print(
            "SHOW SUMMARY",
            row,
            col
        )

        owner = self.owners[row]

        holdings = self.matrix.get(
            owner,
            {}
        )

        lines = []

        lines.append(
            f"Owner : {owner}"
        )

        lines.append(
            "-" * 40
        )

        for khewat, share in sorted(
            holdings.items()
        ):

            lines.append(
                f"{khewat}    :    {share}"
            )

        lines.append("")

        lines.append(
            f"Total Holdings : "
            f"{len(holdings)}"
        )

        QMessageBox.information(
            self,
            f"Owner Summary - {owner}",
            "\n".join(lines)
        )